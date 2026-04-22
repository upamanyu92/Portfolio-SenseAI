from io import BytesIO
from typing import Any, Dict, List, Optional

import openpyxl
import pypdf
from docx import Document


def _normalize_header(value: Any) -> str:
    return str(value or "").strip().lower().replace("\n", " ")


def _to_float(value: Any) -> Optional[float]:
    if value is None or value == "":
        return None
    if isinstance(value, (int, float)):
        return float(value)

    text = str(value).strip().replace(",", "")
    if text.endswith("%"):
        return None

    try:
        return float(text)
    except ValueError:
        return None


def _index_of(headers: Dict[str, int], aliases: List[str]) -> Optional[int]:
    for alias in aliases:
        idx = headers.get(alias)
        if idx is not None:
            return idx
    return None


def _cell(row: tuple[Any, ...], idx: Optional[int]) -> Any:
    if idx is None or idx >= len(row):
        return None
    return row[idx]


def extract_holdings(file_bytes: bytes, extension: str) -> List[Dict[str, Any]]:
    extension = extension.lower()
    if extension != ".xlsx":
        return []

    workbook = openpyxl.load_workbook(BytesIO(file_bytes), data_only=True)
    holdings: List[Dict[str, Any]] = []

    for sheet in workbook.worksheets:
        header_map: Dict[str, int] | None = None
        instrument_type = "stock"

        for row in sheet.iter_rows(values_only=True):
            values = [str(cell or "").strip() for cell in row]
            normalized = [_normalize_header(cell) for cell in row]

            if header_map is None:
                candidate = {name: idx for idx, name in enumerate(normalized) if name}
                has_name = any(key in candidate for key in ["stock name", "scheme name", "name"])
                has_qty = any(key in candidate for key in ["quantity", "qty", "units"])
                if has_name and has_qty:
                    header_map = candidate
                    if "scheme name" in candidate or "folio no." in candidate or "folio no" in candidate:
                        instrument_type = "mutual_fund"
                continue

            if not any(values):
                continue

            name_idx = _index_of(header_map, ["stock name", "scheme name", "name"])
            qty_idx = _index_of(header_map, ["quantity", "qty", "units"])
            isin_idx = _index_of(header_map, ["isin"])
            category_idx = _index_of(header_map, ["category", "sub-category", "sub category"])
            invested_idx = _index_of(header_map, ["invested value", "buy value"])
            current_idx = _index_of(header_map, ["current value", "closing value"])
            source_idx = _index_of(header_map, ["source"])

            name = str(_cell(row, name_idx) or "").strip()
            quantity = _to_float(_cell(row, qty_idx))
            if not name or quantity is None:
                continue

            holding: Dict[str, Any] = {
                "name": name,
                "instrument_type": instrument_type,
                "quantity": quantity,
            }

            isin = str(_cell(row, isin_idx) or "").strip()
            if isin:
                holding["isin"] = isin

            category = str(_cell(row, category_idx) or "").strip()
            if category:
                holding["category"] = category

            invested_value = _to_float(_cell(row, invested_idx))
            if invested_value is not None:
                holding["invested_value"] = invested_value

            current_value = _to_float(_cell(row, current_idx))
            if current_value is not None:
                holding["current_value"] = current_value

            source = str(_cell(row, source_idx) or "").strip()
            if source:
                holding["source"] = source

            holdings.append(holding)

    return holdings


def extract_text(file_bytes: bytes, extension: str) -> str:
    extension = extension.lower()

    if extension == ".pdf":
        reader = pypdf.PdfReader(BytesIO(file_bytes))
        return " ".join((page.extract_text() or "") for page in reader.pages).strip()

    if extension == ".xlsx":
        workbook = openpyxl.load_workbook(BytesIO(file_bytes), data_only=True)
        sheet = workbook.active
        rows = []
        for row in sheet.iter_rows(values_only=True):
            rows.append(",".join("" if cell is None else str(cell) for cell in row))
        return "\n".join(rows).strip()

    if extension == ".docx":
        document = Document(BytesIO(file_bytes))
        return "\n".join(paragraph.text for paragraph in document.paragraphs).strip()

    return ""
